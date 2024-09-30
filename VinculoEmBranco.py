import openpyxl 

wb = openpyxl.load_workbook('consolidado.xlsx')


ws = wb.active
ws = wb['query (1)']


vinculo_col = None
nome_col = None
status_col = None
cpf_col = None  


for col in ws.iter_cols(1, ws.max_column):
    if col[0].value == 'Vínculo':
        vinculo_col = col[0].column  
    elif col[0].value == 'Nome':
        nome_col = col[0].column  
    elif col[0].value == 'Status Contratação':
        status_col = col[0].column  
    elif col[0].value == 'CPF':  
        cpf_col = col[0].column  


if vinculo_col is None or nome_col is None or status_col is None or cpf_col is None:
    raise ValueError("Coluna 'Vínculo', 'Nome', 'Status Contratação' ou 'CPF' não encontrada.")


for row in range(2, ws.max_row + 1):  
    vinculo_value = ws.cell(row=row, column=vinculo_col).value
    status_value = ws.cell(row=row, column=status_col).value
    

    if status_value != 'INATIVO':
        if vinculo_value is None or vinculo_value == '#N/D':
            nome_value = ws.cell(row=row, column=nome_col).value
            cpf_value = ws.cell(row=row, column=cpf_col).value 
            print(f'Linha {row}: Nome = {nome_value}, CPF = {cpf_value}')  
