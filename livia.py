import openpyxl 

wb = openpyxl.load_workbook('CENSO DGPE_Final.xlsm')

ws = wb.active

ws = wb['Base tratada']

print('Total number of rows: ' + str(ws.max_row) + '. And total number of columns: ' + str(ws.max_column))
